from typing import Optional, List, Tuple

import openpyxl
from reports.base_report import BaseReport


class TeacherAttendanceReport(BaseReport):
    key = "attendance"
    title = "Отчет по посещаемости преподавателей (< 40%)"

    def __init__(self, threshold_percent: float = 40.0):
        self.threshold = threshold_percent

    def _find_col_exact(self, ws, header_name: str) -> Optional[int]:
        target = header_name.strip().lower()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if isinstance(v, str) and v.strip().lower() == target:
                return c
        return None


    def build(self, wb: openpyxl.Workbook) -> str:
        ws = wb.active

        teacher_col = self._find_col_exact(ws, "ФИО преподавателя")
        attend_col = self._find_col_exact(ws, "Средняя посещаемость")

        if not teacher_col or not attend_col:
            return (
                "❌ Не нашёл нужные колонки.\n"
                "Ожидаю заголовки: «ФИО преподавателя», «Средняя посещаемость»."
            )

        bad: List[Tuple[str, float]] = []

        for r in range(2, ws.max_row + 1):
            teacher = ws.cell(r, teacher_col).value
            attend = ws.cell(r, attend_col).value

            if not teacher or attend is None:
                continue

            teacher = str(teacher).strip()
            p = self._to_percent(attend)
            if not teacher or p is None:
                continue

            if p < self.threshold:
                bad.append((teacher, p))

        bad.sort(key=lambda x: x[1])

        if not bad:
            return f"✅ Нет преподавателей с посещаемостью ниже {self.threshold:g}%."

        lines = [f"⚠️ Преподаватели с посещаемостью ниже {self.threshold:g}%", ""]
        for i, (t, p) in enumerate(bad, start=1):
            lines.append(f"{i}. {t} — {p:.1f}%")

        return "\n".join(lines)
