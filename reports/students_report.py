from typing import Optional, List, Tuple
import openpyxl
from reports.base_report import BaseReport


class StudentsRiskReport(BaseReport):
    key = "students"
    title = "Отчет по студентам"

    def _find_col_exact(self, ws, header_name: str) -> Optional[int]:
        target = header_name.strip().lower()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if isinstance(v, str) and v.strip().lower() == target:
                return c
        return None

    def _to_float(self, v) -> Optional[float]:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            s = v.strip().replace(",", ".")
            if not s:
                return None
            try:
                return float(s)
            except ValueError:
                return None
        return None

    def build(self, wb: openpyxl.Workbook) -> str:
        ws = wb.active

        name_col = self._find_col_exact(ws, "FIO")
        hw_col = self._find_col_exact(ws, "Homework")
        cw_col = self._find_col_exact(ws, "Classroom")

        if not name_col or not hw_col or not cw_col:
            return (
                "❌ Не нашёл нужные столбцы по точным заголовкам.\n"
                "Ожидаю в первой строке: FIO, Homework, Classroom\n"
                f"Найдено: FIO={name_col}, Homework={hw_col}, Classroom={cw_col}"
            )

        both: List[Tuple[str, float, float]] = []
        only_hw: List[Tuple[str, float, float]] = []
        only_cw: List[Tuple[str, float, float]] = []

        for r in range(2, ws.max_row + 1):
            fio_val = ws.cell(r, name_col).value
            if fio_val is None:
                continue
            fio = str(fio_val).strip()
            if not fio:
                continue

            hw = self._to_float(ws.cell(r, hw_col).value)
            cw = self._to_float(ws.cell(r, cw_col).value)
            if hw is None or cw is None:
                continue

            hw_is_1 = abs(hw - 1.0) < 1e-9
            cw_lt_3 = cw < 3.0

            if hw_is_1 and cw_lt_3:
                both.append((fio, hw, cw))
            if hw_is_1:
                only_hw.append((fio, hw, cw))
            if cw_lt_3:
                only_cw.append((fio, hw, cw))

        def format_block(title: str, items: List[Tuple[str, float, float]], limit: int = 80) -> str:
            if not items:
                return f"{title}\n— нет"
            lines = [title]
            shown = items[:limit]
            for i, (fio, hw, cw) in enumerate(shown, start=1):
                lines.append(f"{i}. {fio} (Homework={hw:g}, Classroom={cw:g})")
            if len(items) > limit:
                lines.append(f"…и ещё {len(items) - limit}")
            return "\n".join(lines)

        parts = [
            "👩‍🎓 Отчёт по студентам",
            "",
            format_block("📌 1) Homework = 1 И Classroom < 3:", both),
            "",
            format_block("📌 2) Homework = 1 (независимо от Classroom):", only_hw),
            "",
            format_block("📌 3) Classroom < 3 (независимо от Homework):", only_cw),
        ]

        text = "\n".join(parts)

        if len(text) > 3800:
            text = text[:3800] + "\n…(сообщение обрезано, слишком длинно)"
        return text
