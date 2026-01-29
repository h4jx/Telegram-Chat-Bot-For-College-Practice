import re
from collections import Counter, defaultdict
from typing import Dict, List, Tuple

import openpyxl

from reports.base_report import BaseReport


class ScheduleCountReport(BaseReport):
    key = "schedule"
    title = "Отчет по выставленному расписанию"

    SUBJECT_RE = re.compile(r"Предмет:\s*(.+)")

    def _find_lesson_columns(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> List[int]:

        cols: List[int] = []
        for c in range(1, ws.max_column + 1):
            header = ws.cell(1, c).value
            if not isinstance(header, str):
                continue
            header = header.strip()
            if header in ("Группа", "Пара", "Время"):
                continue
            if header == "Время":
                continue
            cols.append(c)
        return cols

    def _extract_subject(self, cell_value: str) -> str | None:

        if not isinstance(cell_value, str):
            return None
        if "Предмет:" not in cell_value:
            return None

        m = self.SUBJECT_RE.search(cell_value)
        if not m:
            return None
        return m.group(1).strip() or None

    def build(self, wb: openpyxl.Workbook) -> str:
        ws = wb.active
        lesson_cols = self._find_lesson_columns(ws)
        if not lesson_cols:
            return "Не смог найти колонки с занятиями (дни недели) в первой строке файла."

        group_counts: Dict[str, Counter] = defaultdict(Counter)

        groups_order: List[str] = []
        seen_groups = set()

        for r in range(2, ws.max_row + 1):
            group_val = ws.cell(r, 1).value
            if group_val is None:
                continue
            group = str(group_val).strip()
            if not group:
                continue

            if group not in seen_groups:
                seen_groups.add(group)
                groups_order.append(group)

            for c in lesson_cols:
                val = ws.cell(r, c).value
                subj = self._extract_subject(val) if isinstance(val, str) else None
                if subj:
                    group_counts[group][subj] += 1

        if not groups_order:
            return "В файле не найдено ни одной группы (колонка A)."

        lines: List[str] = []
        for group in groups_order:
            counts = group_counts.get(group, Counter())
            lines.append(f"📌 Группа: {group}")

            if not counts:
                lines.append("Занятий (с 'Предмет: ...') не найдено.")
                lines.append("")
                continue

            total = sum(counts.values())
            lines.append(f"Всего пар: {total}")
            lines.append("Количество пар по дисциплинам:")

            for subject, n in counts.most_common():
                lines.append(f"• {subject} — {n}")

            lines.append("")

        result = "\n".join(lines).strip()
        if len(result) > 3800:
            result = result[:3800] + "\n…(сообщение обрезано, слишком много данных)"

        return result
