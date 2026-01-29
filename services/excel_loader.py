from io import BytesIO
from pathlib import Path

import openpyxl
import xlrd
from aiogram import Bot
from aiogram.types import Message


class ExcelLoader:
    async def load_workbook_from_message(self, message: Message, bot: Bot) -> openpyxl.Workbook:
        if not message.document:
            raise ValueError("Нет документа в сообщении")

        filename = message.document.file_name or "file"
        ext = Path(filename).suffix.lower()

        if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"):
            raise ValueError("Нужен Excel-файл: .xls или .xlsx")

        buf = BytesIO()
        await bot.download(message.document, destination=buf)
        data = buf.getvalue()

        # .xlsx / .xlsm
        if ext != ".xls":
            buf2 = BytesIO(data)
            buf2.seek(0)
            return openpyxl.load_workbook(buf2, data_only=True)

        return self._xls_to_openpyxl(data)

    def _xls_to_openpyxl(self, data: bytes) -> openpyxl.Workbook:
        book = xlrd.open_workbook(file_contents=data)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for i in range(book.nsheets):
            sh = book.sheet_by_index(i)
            ws = wb.create_sheet(title=sh.name[:31])

            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    val = sh.cell_value(r, c)
                    ws.cell(row=r + 1, column=c + 1, value=val)

        return wb
